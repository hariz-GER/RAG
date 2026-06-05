import os
from pypdf import PdfReader
import openpyxl
from docx import Document


def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    text = ""
    try:
        reader = PdfReader(file_path)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    except Exception as e:
        print(f"Error reading PDF: {e}")
    return text


def extract_text_from_excel(file_path):
    """Extract text from Excel file"""
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_path)
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            text += f"=== Sheet: {sheet} ===\n"
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        text += str(cell) + " "
                text += "\n"
    except Exception as e:
        print(f"Error reading Excel: {e}")
    return text


def extract_text_from_word(file_path):
    """Extract text from Word document"""
    text = ""
    try:
        doc = Document(file_path)
        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"
    except Exception as e:
        print(f"Error reading Word: {e}")
    return text


def process_file(file_path):
    """Process any supported file type and extract text"""
    file_extension = os.path.splitext(file_path)[1].lower()
    
    if file_extension == ".pdf":
        return extract_text_from_pdf(file_path)
    elif file_extension in [".xlsx", ".xls"]:
        return extract_text_from_excel(file_path)
    elif file_extension in [".docx", ".doc"]:
        return extract_text_from_word(file_path)
    else:
        raise ValueError(f"Unsupported file type: {file_extension}")
